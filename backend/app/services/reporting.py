"""
Reporting Service - PDF/Excel Export
Monthly summaries and automated insights
"""

from datetime import datetime, timedelta
from typing import Dict, List, Optional
from collections import Counter
import os
import io
import logging

from app.services.analytics import get_analytics_service
from app.services.admin_analytics import get_admin_analytics

logger = logging.getLogger("REPORTING")

# Optional imports for PDF/Excel
try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import inch, cm
    from reportlab.lib.colors import HexColor
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle
    REPORTLAB_AVAILABLE = True
except:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Fill, PatternFill, Alignment
    from openpyxl.chart import BarChart, Reference
    OPENPYXL_AVAILABLE = True
except:
    OPENPYXL_AVAILABLE = False


class ReportingService:
    """
    Generate monthly reports in PDF/Excel
    """
    
    def __init__(self):
        self.output_dir = "/tmp/doer_reports"
        os.makedirs(self.output_dir, exist_ok=True)
        self.analytics = get_analytics_service()
        self.admin = get_admin_analytics()
    
    def generate_monthly_summary(self, month: int = None, year: int = None) -> Dict:
        """Generate monthly case summary data"""
        now = datetime.now()
        month = month or now.month
        year = year or now.year
        
        cases = self.analytics.demo_cases
        
        # Filter cases for the month
        month_cases = []
        for c in cases:
            case_date = datetime.fromisoformat(c["started_at"])
            if case_date.month == month and case_date.year == year:
                month_cases.append(c)
        
        # Calculate metrics
        resolved = [c for c in month_cases if c["phase"] == "closed"]
        
        type_dist = Counter(c["type"] for c in month_cases)
        state_dist = Counter(c["state"] for c in month_cases)
        
        # Resolution times
        resolution_times = []
        for c in resolved:
            if c.get("resolved_at"):
                start = datetime.fromisoformat(c["started_at"])
                end = datetime.fromisoformat(c["resolved_at"])
                resolution_times.append((end - start).days)
        
        avg_resolution = sum(resolution_times) / len(resolution_times) if resolution_times else 0
        
        # NPS for the month
        nps = self.analytics.calculate_nps()
        
        return {
            "period": f"{year}-{month:02d}",
            "month_name": datetime(year, month, 1).strftime("%B %Y"),
            "total_cases_filed": len(month_cases),
            "cases_resolved": len(resolved),
            "active_cases": len(month_cases) - len(resolved),
            "avg_resolution_days": round(avg_resolution, 1),
            "case_type_distribution": dict(type_dist),
            "state_distribution": dict(state_dist),
            "nps_score": nps["nps"],
            "ai_resolved": sum(1 for c in resolved if c.get("resolved_by") == "ai"),
            "human_resolved": sum(1 for c in resolved if c.get("resolved_by") == "human"),
        }
    
    def export_pdf_report(self, month: int = None, year: int = None) -> str:
        """Export monthly report as PDF"""
        summary = self.generate_monthly_summary(month, year)
        insights = self.admin.generate_automated_insights()
        funnel = self.admin.get_cases_by_status_funnel()
        
        if not REPORTLAB_AVAILABLE:
            logger.warning("ReportLab not available, simulating PDF")
            filename = f"{self.output_dir}/report_{summary['period']}_simulated.pdf"
            with open(filename, 'w') as f:
                f.write(f"SIMULATED PDF: {summary}")
            return filename
        
        filename = f"{self.output_dir}/monthly_report_{summary['period']}.pdf"
        c = canvas.Canvas(filename, pagesize=A4)
        width, height = A4
        
        # Header
        c.setFillColor(HexColor("#1e40af"))
        c.rect(0, height - 80, width, 80, fill=True)
        c.setFillColor(HexColor("#ffffff"))
        c.setFont("Helvetica-Bold", 24)
        c.drawCentredString(width/2, height - 45, "DOER Platform")
        c.setFont("Helvetica", 14)
        c.drawCentredString(width/2, height - 65, f"Monthly Report - {summary['month_name']}")
        
        y = height - 120
        
        # Summary boxes
        c.setFillColor(HexColor("#1f2937"))
        c.setFont("Helvetica-Bold", 14)
        c.drawString(50, y, "Executive Summary")
        
        y -= 30
        metrics = [
            ("Cases Filed", summary["total_cases_filed"]),
            ("Cases Resolved", summary["cases_resolved"]),
            ("Avg Resolution", f"{summary['avg_resolution_days']} days"),
            ("NPS Score", f"{summary['nps_score']:.0f}"),
        ]
        
        x = 50
        for label, value in metrics:
            c.setFillColor(HexColor("#e5e7eb"))
            c.roundRect(x, y - 50, 120, 55, 5, fill=True)
            c.setFillColor(HexColor("#1f2937"))
            c.setFont("Helvetica-Bold", 20)
            c.drawCentredString(x + 60, y - 25, str(value))
            c.setFont("Helvetica", 10)
            c.drawCentredString(x + 60, y - 45, label)
            x += 130
        
        y -= 90
        
        # Case type distribution
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, y, "Case Type Distribution")
        y -= 20
        c.setFont("Helvetica", 10)
        for case_type, count in summary["case_type_distribution"].items():
            c.drawString(70, y, f"• {case_type.replace('_', ' ').title()}: {count} cases")
            y -= 15
        
        y -= 20
        
        # Insights
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, y, "Automated Insights")
        y -= 20
        c.setFont("Helvetica", 10)
        for insight in insights[:5]:
            icon = "⚠️" if insight["type"] == "warning" else "✅"
            c.drawString(70, y, f"{icon} {insight['message']}")
            y -= 15
        
        # Footer
        c.setFillColor(HexColor("#6b7280"))
        c.setFont("Helvetica", 8)
        c.drawCentredString(width/2, 30, f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        
        c.save()
        logger.info(f"📊 PDF report generated: {filename}")
        return filename
    
    def export_excel_report(self, month: int = None, year: int = None) -> str:
        """Export monthly report as Excel"""
        summary = self.generate_monthly_summary(month, year)
        funnel = self.admin.get_cases_by_status_funnel()
        leaderboard = self.admin.get_talent_leaderboard()
        
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available, simulating Excel")
            filename = f"{self.output_dir}/report_{summary['period']}_simulated.xlsx"
            with open(filename, 'w') as f:
                f.write(f"SIMULATED EXCEL: {summary}")
            return filename
        
        filename = f"{self.output_dir}/monthly_report_{summary['period']}.xlsx"
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Header
        ws_summary["A1"] = "DOER Platform Monthly Report"
        ws_summary["A1"].font = Font(size=16, bold=True)
        ws_summary["A2"] = summary["month_name"]
        
        # Metrics
        row = 4
        metrics = [
            ("Total Cases Filed", summary["total_cases_filed"]),
            ("Cases Resolved", summary["cases_resolved"]),
            ("Active Cases", summary["active_cases"]),
            ("Avg Resolution Days", summary["avg_resolution_days"]),
            ("NPS Score", summary["nps_score"]),
            ("AI Resolved", summary["ai_resolved"]),
            ("Human Resolved", summary["human_resolved"]),
        ]
        
        for label, value in metrics:
            ws_summary[f"A{row}"] = label
            ws_summary[f"B{row}"] = value
            row += 1
        
        # Case types sheet
        ws_types = wb.create_sheet("Case Types")
        ws_types["A1"] = "Case Type"
        ws_types["B1"] = "Count"
        row = 2
        for case_type, count in summary["case_type_distribution"].items():
            ws_types[f"A{row}"] = case_type.replace("_", " ").title()
            ws_types[f"B{row}"] = count
            row += 1
        
        # Funnel sheet
        ws_funnel = wb.create_sheet("Status Funnel")
        ws_funnel["A1"] = "Phase"
        ws_funnel["B1"] = "Count"
        ws_funnel["C1"] = "Percentage"
        row = 2
        for phase_data in funnel["funnel"]:
            ws_funnel[f"A{row}"] = phase_data["label"]
            ws_funnel[f"B{row}"] = phase_data["count"]
            ws_funnel[f"C{row}"] = f"{phase_data['percentage']}%"
            row += 1
        
        # Leaderboard sheet
        ws_lead = wb.create_sheet("Talent Leaderboard")
        headers = ["Rank", "Name", "Cases", "Resolved", "Avg Days", "Rating", "Score"]
        for i, h in enumerate(headers, 1):
            ws_lead.cell(row=1, column=i, value=h)
        
        for rank, talent in enumerate(leaderboard["leaderboard"], 1):
            ws_lead.cell(row=rank+1, column=1, value=rank)
            ws_lead.cell(row=rank+1, column=2, value=talent["name"])
            ws_lead.cell(row=rank+1, column=3, value=talent["cases_handled"])
            ws_lead.cell(row=rank+1, column=4, value=talent["cases_resolved"])
            ws_lead.cell(row=rank+1, column=5, value=talent["avg_resolution_days"])
            ws_lead.cell(row=rank+1, column=6, value=talent["avg_rating"])
            ws_lead.cell(row=rank+1, column=7, value=talent["score"])
        
        wb.save(filename)
        logger.info(f"📊 Excel report generated: {filename}")
        return filename
    
    def get_report_files(self) -> List[Dict]:
        """List available reports"""
        files = []
        for f in os.listdir(self.output_dir):
            filepath = os.path.join(self.output_dir, f)
            files.append({
                "filename": f,
                "filepath": filepath,
                "size_kb": os.path.getsize(filepath) / 1024,
                "created": datetime.fromtimestamp(os.path.getctime(filepath)).isoformat(),
            })
        return files


# Singleton
_reporting_service: Optional[ReportingService] = None

def get_reporting_service() -> ReportingService:
    global _reporting_service
    if _reporting_service is None:
        _reporting_service = ReportingService()
    return _reporting_service
